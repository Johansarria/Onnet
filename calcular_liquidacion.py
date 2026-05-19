import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def get_months_by_year(cosecha_date, target_date=pd.to_datetime('2025-06-01')):
    if pd.isna(cosecha_date):
        return {2024: 0, 2025: 0}
    cosecha_date = pd.to_datetime(cosecha_date)
    months = pd.date_range(start=cosecha_date, end=target_date, freq='MS')
    years_count = {2024: 0, 2025: 0}
    for m in months:
        if m.year in years_count:
            years_count[m.year] += 1
    return years_count

def normalize_key(s):
    import unicodedata
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('utf-8')
    return s.strip().lower()

def main():
    excel_file = "Prueba Consultor contratos e Inventario de red (1).xlsx"
    out_file = "Liquidacion_Junio_2025.xlsx"
    print(f"Leyendo archivo excel: {excel_file}")
    
    xls = pd.ExcelFile(excel_file)
    sheet_names = xls.sheet_names
    
    sheet_inv = [s for s in sheet_names if 'inventario' in s.lower()][0]
    sheet_liq = [s for s in sheet_names if 'liquidac' in s.lower()][0]
    sheet_pre = [s for s in sheet_names if 'precio' in s.lower()][0]
    
    # 1. Cargar Precios
    df_precios_raw = pd.read_excel(excel_file, sheet_name=sheet_pre)
    
    prices_raw = {
        'Electrificadora': {
            'alto,moderado,incipiente': {2024: {}, 2025: {}},
            'bajo,limitado': {2024: {}, 2025: {}}
        },
        'Telecomunicaciones': {
            'alto,moderado,incipiente': {2024: {}, 2025: {}},
            'bajo,limitado': {2024: {}, 2025: {}}
        }
    }
    
    # Electrificadora (Filas 1 a 6)
    for idx in range(1, 7):
        name = normalize_key(df_precios_raw.iloc[idx, 0])
        prices_raw['Electrificadora']['alto,moderado,incipiente'][2024][name] = float(df_precios_raw.iloc[idx, 2])
        prices_raw['Electrificadora']['bajo,limitado'][2024][name] = float(df_precios_raw.iloc[idx, 3])
        prices_raw['Electrificadora']['alto,moderado,incipiente'][2025][name] = float(df_precios_raw.iloc[idx, 4])
        prices_raw['Electrificadora']['bajo,limitado'][2025][name] = float(df_precios_raw.iloc[idx, 5])
        
    # Telecomunicaciones (Filas 10 a 14)
    for idx in range(10, 15):
        name = normalize_key(df_precios_raw.iloc[idx, 0])
        prices_raw['Telecomunicaciones']['alto,moderado,incipiente'][2024][name] = float(df_precios_raw.iloc[idx, 2])
        prices_raw['Telecomunicaciones']['bajo,limitado'][2024][name] = float(df_precios_raw.iloc[idx, 3])
        prices_raw['Telecomunicaciones']['alto,moderado,incipiente'][2025][name] = float(df_precios_raw.iloc[idx, 4])
        prices_raw['Telecomunicaciones']['bajo,limitado'][2025][name] = float(df_precios_raw.iloc[idx, 5])
        
    # Poste/Torre en Telecomunicaciones no existe, lo seteamos en 0.0
    for perf in ['alto,moderado,incipiente', 'bajo,limitado']:
        for year in [2024, 2025]:
            prices_raw['Telecomunicaciones'][perf][year]['poste/torre'] = 0.0
            
    # 2. Cargar Inventario y OTs a Liquidar
    df_liq = pd.read_excel(excel_file, sheet_name=sheet_liq, header=1).dropna(subset=['OT'])
    df_inv = pd.read_excel(excel_file, sheet_name=sheet_inv).dropna(subset=['OT_ORIGINAL'])
    
    # Filtrar el inventario para procesar únicamente las OTs de liquidación
    df_inv_filtered = df_inv[df_inv['OT_ORIGINAL'].isin(df_liq['OT'])].copy()
    
    # Asegurar que todas las cantidades de postes y ductos sean numéricas
    height_cols = [8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 20, 21]
    all_qty_cols = height_cols + ['Torre', 'Q Ductos']
    for c in all_qty_cols:
        df_inv_filtered[c] = pd.to_numeric(df_inv_filtered[c], errors='coerce').fillna(0.0)
        
    # Clasificación de desempeño
    def get_perf_key(desempeno):
        if not isinstance(desempeno, str):
            return 'alto,moderado,incipiente'
        d = desempeno.lower().strip()
        if d in ['alto', 'moderado', 'incipiente']:
            return 'alto,moderado,incipiente'
        elif d in ['bajo', 'limitado']:
            return 'bajo,limitado'
        else:
            return 'alto,moderado,incipiente'
            
    # Acumuladores globales para el Resumen
    v_le8_2024_tot = 0.0
    v_le8_2025_tot = 0.0
    v_8_10_2024_tot = 0.0
    v_8_10_2025_tot = 0.0
    v_gt10_2024_tot = 0.0
    v_gt10_2025_tot = 0.0
    v_torre_2024_tot = 0.0
    v_torre_2025_tot = 0.0
    v_duct1_2024_tot = 0.0
    v_duct1_2025_tot = 0.0
    v_duct2_2024_tot = 0.0
    v_duct2_2025_tot = 0.0

    # Calcular costos e información agregada por OT
    ot_data = {}
    
    for _, row in df_inv_filtered.iterrows():
        ot = row['OT_ORIGINAL']
        empresa = row['Electrificacora/Telco']
        municipio = row['Municipio']
        cod_mun = row['Código municipio']
        desempeno = row['Desempeño']
        cosecha = row['Cosecha']
        
        emp_key = 'Electrificadora' if 'electri' in str(empresa).lower() else 'Telecomunicaciones'
        perf_key = get_perf_key(desempeno)
        
        # Meses de servicio
        months = get_months_by_year(cosecha)
        m_2024 = months[2024]
        m_2025 = months[2025]
        
        # Cantidades en esta fila
        q_le8 = float(row[8])
        q_8_10 = float(row[9] + row[10])
        q_gt10 = float(sum(row[h] for h in [11, 12, 13, 14, 15, 16, 17, 18, 20, 21]))
        q_torre = float(row['Torre'])
        q_ductos = float(row['Q Ductos'])
        
        # Mapeo a las columnas del formato de liquidación (8, 10, 12, 14, 16, 18, 21, Torre)
        q_mapped_8 = float(row[8])
        q_mapped_10 = float(row[9] + row[10])
        q_mapped_12 = float(row[11] + row[12])
        q_mapped_14 = float(row[13] + row[14])
        q_mapped_16 = float(row[15] + row[16])
        q_mapped_18 = float(row[17] + row[18])
        q_mapped_21 = float(row[20] + row[21])
        q_mapped_torre = float(row['Torre'])
        q_mapped_total = q_mapped_8 + q_mapped_10 + q_mapped_12 + q_mapped_14 + q_mapped_16 + q_mapped_18 + q_mapped_21 + q_mapped_torre
        
        # Precios
        p_le8_2024 = prices_raw[emp_key][perf_key][2024]['poste menor o igual a 8m']
        p_le8_2025 = prices_raw[emp_key][perf_key][2025]['poste menor o igual a 8m']
        
        p_8_10_2024 = prices_raw[emp_key][perf_key][2024]['poste mayor a 8m menor o igual a 10m']
        p_8_10_2025 = prices_raw[emp_key][perf_key][2025]['poste mayor a 8m menor o igual a 10m']
        
        p_gt10_2024 = prices_raw[emp_key][perf_key][2024]['poste mayor a 10m']
        p_gt10_2025 = prices_raw[emp_key][perf_key][2025]['poste mayor a 10m']
        
        p_torre_2024 = prices_raw[emp_key][perf_key][2024]['poste/torre']
        p_torre_2025 = prices_raw[emp_key][perf_key][2025]['poste/torre']
        
        # Ductos
        p_duct1_2024 = prices_raw[emp_key][perf_key][2024]['1 ducto de comparticion']
        p_duct1_2025 = prices_raw[emp_key][perf_key][2025]['1 ducto de comparticion']
        p_duct2_2024 = prices_raw[emp_key][perf_key][2024]['2 ductos de comparticion']
        p_duct2_2025 = prices_raw[emp_key][perf_key][2025]['2 ductos de comparticion']
        
        # Cálculos de valor para esta fila
        val_le8 = q_le8 * (m_2024 * p_le8_2024 + m_2025 * p_le8_2025)
        val_8_10 = q_8_10 * (m_2024 * p_8_10_2024 + m_2025 * p_8_10_2025)
        val_gt10 = q_gt10 * (m_2024 * p_gt10_2024 + m_2025 * p_gt10_2025)
        val_torre = q_torre * (m_2024 * p_torre_2024 + m_2025 * p_torre_2025)
        
        val_duct1 = q_ductos * (m_2024 * p_duct1_2024 + m_2025 * p_duct1_2025)
        val_duct2 = q_ductos * (m_2024 * p_duct2_2024 + m_2025 * p_duct2_2025)
        
        val_postes_tot = val_le8 + val_8_10 + val_gt10 + val_torre
        
        # Acumular globales
        v_le8_2024_tot += q_le8 * m_2024 * p_le8_2024
        v_le8_2025_tot += q_le8 * m_2025 * p_le8_2025
        v_8_10_2024_tot += q_8_10 * m_2024 * p_8_10_2024
        v_8_10_2025_tot += q_8_10 * m_2025 * p_8_10_2025
        v_gt10_2024_tot += q_gt10 * m_2024 * p_gt10_2024
        v_gt10_2025_tot += q_gt10 * m_2025 * p_gt10_2025
        v_torre_2024_tot += q_torre * m_2024 * p_torre_2024
        v_torre_2025_tot += q_torre * m_2025 * p_torre_2025
        v_duct1_2024_tot += q_ductos * m_2024 * p_duct1_2024
        v_duct1_2025_tot += q_ductos * m_2025 * p_duct1_2025
        v_duct2_2024_tot += q_ductos * m_2024 * p_duct2_2024
        v_duct2_2025_tot += q_ductos * m_2025 * p_duct2_2025

        if ot not in ot_data:
            ot_data[ot] = {
                'Municipio': municipio,
                'Codigo_Municipio': cod_mun,
                'Cosecha': cosecha,
                'Q_8': 0.0,
                'Q_10': 0.0,
                'Q_12': 0.0,
                'Q_14': 0.0,
                'Q_16': 0.0,
                'Q_18': 0.0,
                'Q_21': 0.0,
                'Q_Torre': 0.0,
                'Q_Total_Postes': 0.0,
                'Q_Ductos': 0.0,
                'Valor_Postes': 0.0,
                'Valor_Ductos_S1': 0.0,
                'Valor_Ductos_S2': 0.0,
                'Total_S1': 0.0,
                'Total_S2': 0.0,
                'Cosecha_earliest': cosecha
            }
            
        # Acumular por OT
        ot_data[ot]['Q_8'] += q_mapped_8
        ot_data[ot]['Q_10'] += q_mapped_10
        ot_data[ot]['Q_12'] += q_mapped_12
        ot_data[ot]['Q_14'] += q_mapped_14
        ot_data[ot]['Q_16'] += q_mapped_16
        ot_data[ot]['Q_18'] += q_mapped_18
        ot_data[ot]['Q_21'] += q_mapped_21
        ot_data[ot]['Q_Torre'] += q_mapped_torre
        ot_data[ot]['Q_Total_Postes'] += q_mapped_total
        ot_data[ot]['Q_Ductos'] += q_ductos
        
        ot_data[ot]['Valor_Postes'] += val_postes_tot
        ot_data[ot]['Valor_Ductos_S1'] += val_duct1
        ot_data[ot]['Valor_Ductos_S2'] += val_duct2
        ot_data[ot]['Total_S1'] += (val_postes_tot + val_duct1)
        ot_data[ot]['Total_S2'] += (val_postes_tot + val_duct2)
        
        # Mantener fecha de cosecha más antigua
        if pd.notna(cosecha) and (pd.isna(ot_data[ot]['Cosecha_earliest']) or cosecha < ot_data[ot]['Cosecha_earliest']):
            ot_data[ot]['Cosecha_earliest'] = cosecha

    print(f"Cálculos realizados para {len(ot_data)} OTs del inventario.")

    # 3. Cargar el archivo original con openpyxl para no destruir formatos ni otras hojas
    wb = openpyxl.load_workbook(excel_file)
    ws_liq = wb[sheet_liq]
    
    # Cabeceras para nuevas columnas
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    cell_font = Font(name=font_family, size=10)
    total_font = Font(name=font_family, size=10, bold=True)
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom_border = Border(top=thin_border_side, bottom=Side(border_style="double", color="000000"))
    subtotal_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    new_cols = [
        (17, "Q Ductos"),
        (18, "Valor Postes"),
        (19, "Valor Ductos S1 (1 Ducto)"),
        (20, "Valor Ductos S2 (2 Ductos)"),
        (21, "Total S1 (1 Ducto)"),
        (22, "Total S2 (2 Ductos)")
    ]
    
    for col_idx, text in new_cols:
        cell = ws_liq.cell(row=2, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    ws_liq.row_dimensions[2].height = 28
    
    # Llenar datos de OTs a Liquidar
    row_count = 0
    for r_idx in range(3, 1000):
        ot_cell = ws_liq.cell(row=r_idx, column=1)
        if ot_cell.value is None or str(ot_cell.value).strip() == "":
            break
            
        ot_val = str(ot_cell.value).strip()
        row_count += 1
        
        if ot_val in ot_data:
            data = ot_data[ot_val]
            
            ws_liq.cell(row=r_idx, column=2, value=data['Municipio']).alignment = Alignment(horizontal="left")
            ws_liq.cell(row=r_idx, column=3, value=data['Codigo_Municipio']).alignment = Alignment(horizontal="center")
            
            cosecha_cell = ws_liq.cell(row=r_idx, column=4, value=data['Cosecha_earliest'])
            cosecha_cell.number_format = 'yyyy-mm-dd'
            cosecha_cell.alignment = Alignment(horizontal="center")
            
            ws_liq.cell(row=r_idx, column=8, value=data['Q_8']).number_format = '#,##0.0'
            ws_liq.cell(row=r_idx, column=9, value=data['Q_10']).number_format = '#,##0.0'
            ws_liq.cell(row=r_idx, column=10, value=data['Q_12']).number_format = '#,##0.0'
            ws_liq.cell(row=r_idx, column=11, value=data['Q_14']).number_format = '#,##0.0'
            ws_liq.cell(row=r_idx, column=12, value=data['Q_16']).number_format = '#,##0.0'
            ws_liq.cell(row=r_idx, column=13, value=data['Q_18']).number_format = '#,##0.0'
            ws_liq.cell(row=r_idx, column=14, value=data['Q_21']).number_format = '#,##0.0'
            ws_liq.cell(row=r_idx, column=15, value=data['Q_Torre']).number_format = '#,##0.0'
            ws_liq.cell(row=r_idx, column=16, value=data['Q_Total_Postes']).number_format = '#,##0.0'
            
            ws_liq.cell(row=r_idx, column=17, value=data['Q_Ductos']).number_format = '#,##0.0'
            
            ws_liq.cell(row=r_idx, column=18, value=data['Valor_Postes']).number_format = '$#,##0.00'
            ws_liq.cell(row=r_idx, column=19, value=data['Valor_Ductos_S1']).number_format = '$#,##0.00'
            ws_liq.cell(row=r_idx, column=20, value=data['Valor_Ductos_S2']).number_format = '$#,##0.00'
            ws_liq.cell(row=r_idx, column=21, value=data['Total_S1']).number_format = '$#,##0.00'
            ws_liq.cell(row=r_idx, column=22, value=data['Total_S2']).number_format = '$#,##0.00'
        else:
            print(f"Alerta: OT '{ot_val}' no se encontró en el inventario.")
            
        for c_idx in range(1, 23):
            cell = ws_liq.cell(row=r_idx, column=c_idx)
            cell.font = cell_font
            cell.border = thin_border
            if c_idx in [8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22]:
                cell.alignment = Alignment(horizontal="right")
                
    tot_row = row_count + 3
    ws_liq.cell(row=tot_row, column=1, value="TOTAL GENERAL").font = total_font
    ws_liq.cell(row=tot_row, column=1).fill = subtotal_fill
    ws_liq.cell(row=tot_row, column=1).border = double_bottom_border
    
    for c_idx in range(2, 23):
        cell = ws_liq.cell(row=tot_row, column=c_idx)
        cell.font = total_font
        cell.fill = subtotal_fill
        cell.border = double_bottom_border
        
        col_letter = get_column_letter(c_idx)
        if c_idx in [8, 9, 10, 11, 12, 13, 14, 15, 16, 17]:
            cell.value = f"=SUM({col_letter}3:{col_letter}{tot_row-1})"
            cell.number_format = '#,##0.0'
            cell.alignment = Alignment(horizontal="right")
        elif c_idx in [18, 19, 20, 21, 22]:
            cell.value = f"=SUM({col_letter}3:{col_letter}{tot_row-1})"
            cell.number_format = '$#,##0.00'
            cell.alignment = Alignment(horizontal="right")
            
    for col in ws_liq.columns:
        max_len = 0
        for cell in col:
            val = cell.value
            if val is not None:
                if str(val).startswith('='):
                    max_len = max(max_len, 12)
                else:
                    max_len = max(max_len, len(str(val)))
        col_letter = get_column_letter(col[0].column)
        ws_liq.column_dimensions[col_letter].width = max(max_len + 3, 11)

    print(f"Llenado de hoja de liquidación completado. {row_count} filas procesadas.")

    # 4. Crear la hoja de Resumen / Dashboard
    if 'Resumen_Liquidacion' in wb.sheetnames:
        wb.remove(wb['Resumen_Liquidacion'])
        
    ws_res = wb.create_sheet('Resumen_Liquidacion', 0)
    ws_res.views.sheetView[0].showGridLines = True
    
    title_font = Font(name=font_family, size=16, bold=True, color="1F4E79")
    ws_res.cell(row=2, column=2, value="Resumen Gerencial de Liquidación").font = title_font
    ws_res.cell(row=3, column=2, value="Periodo de Liquidación: Junio 2025 (Retroactivos Aplicados)").font = Font(name=font_family, size=11, italic=True)
    
    ws_res.cell(row=5, column=2, value="ESCENARIO 1: Liquidación con Ductos como '1 Ducto de Compartición'").font = Font(name=font_family, size=11, bold=True, color="1F4E79")
    
    headers_res = ["Categoría de Infraestructura", "Cantidad Total", "Valor Retroactivo 2024", "Valor Periodo 2025", "Valor Total Liquidación"]
    for col_idx, h in enumerate(headers_res, start=2):
        cell = ws_res.cell(row=6, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    # Obtener dataframe acumulado para cantidades del resumen
    df_res_all = pd.DataFrame(ot_data.values())
    
    q_le8_t = df_res_all['Q_8'].sum()
    q_8_10_t = df_res_all['Q_10'].sum()
    q_gt10_t = df_res_all['Q_12'].sum() + df_res_all['Q_14'].sum() + df_res_all['Q_16'].sum() + df_res_all['Q_18'].sum() + df_res_all['Q_21'].sum()
    q_torre_t = df_res_all['Q_Torre'].sum()
    q_duct_t = df_res_all['Q_Ductos'].sum()
    
    row_data_s1 = [
        ("Postes menor o igual a 8m", q_le8_t, v_le8_2024_tot, v_le8_2025_tot, v_le8_2024_tot + v_le8_2025_tot),
        ("Postes mayor a 8m menor o igual a 10m", q_8_10_t, v_8_10_2024_tot, v_8_10_2025_tot, v_8_10_2024_tot + v_8_10_2025_tot),
        ("Postes mayor a 10m", q_gt10_t, v_gt10_2024_tot, v_gt10_2025_tot, v_gt10_2024_tot + v_gt10_2025_tot),
        ("Postes/Torre", q_torre_t, v_torre_2024_tot, v_torre_2025_tot, v_torre_2024_tot + v_torre_2025_tot),
        ("Ductos (Tarifa 1 Ducto)", q_duct_t, v_duct1_2024_tot, v_duct1_2025_tot, v_duct1_2024_tot + v_duct1_2025_tot)
    ]
    
    for r_idx, data in enumerate(row_data_s1, start=7):
        for c_idx, val in enumerate(data, start=2):
            cell = ws_res.cell(row=r_idx, column=c_idx, value=val)
            cell.font = cell_font
            cell.border = thin_border
            if c_idx == 2:
                cell.alignment = Alignment(horizontal="left")
            elif c_idx == 3:
                cell.number_format = '#,##0.0'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="right")
                
    tot_row_s1 = 12
    ws_res.cell(row=tot_row_s1, column=2, value="TOTAL LIQUIDACIÓN (ESCENARIO 1)").font = total_font
    ws_res.cell(row=tot_row_s1, column=2).fill = subtotal_fill
    ws_res.cell(row=tot_row_s1, column=2).border = double_bottom_border
    
    ws_res.cell(row=tot_row_s1, column=3, value=f"=SUM(C7:C11)").font = total_font
    ws_res.cell(row=tot_row_s1, column=3).number_format = '#,##0.0'
    ws_res.cell(row=tot_row_s1, column=3).fill = subtotal_fill
    ws_res.cell(row=tot_row_s1, column=3).border = double_bottom_border
    
    for c_idx, col_let in enumerate(['D', 'E', 'F'], start=4):
        cell = ws_res.cell(row=tot_row_s1, column=c_idx, value=f"=SUM({col_let}7:{col_let}11)")
        cell.font = total_font
        cell.number_format = '$#,##0.00'
        cell.fill = subtotal_fill
        cell.border = double_bottom_border
        
    # Cuadro de Resumen Escenario 2
    ws_res.cell(row=15, column=2, value="ESCENARIO 2: Liquidación con Ductos como '2 Ductos de Compartición'").font = Font(name=font_family, size=11, bold=True, color="1F4E79")
    
    for col_idx, h in enumerate(headers_res, start=2):
        cell = ws_res.cell(row=16, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    row_data_s2 = [
        ("Postes menor o igual a 8m", q_le8_t, v_le8_2024_tot, v_le8_2025_tot, v_le8_2024_tot + v_le8_2025_tot),
        ("Postes mayor a 8m menor o igual a 10m", q_8_10_t, v_8_10_2024_tot, v_8_10_2025_tot, v_8_10_2024_tot + v_8_10_2025_tot),
        ("Postes mayor a 10m", q_gt10_t, v_gt10_2024_tot, v_gt10_2025_tot, v_gt10_2024_tot + v_gt10_2025_tot),
        ("Postes/Torre", q_torre_t, v_torre_2024_tot, v_torre_2025_tot, v_torre_2024_tot + v_torre_2025_tot),
        ("Ductos (Tarifa 2 Ductos)", q_duct_t, v_duct2_2024_tot, v_duct2_2025_tot, v_duct2_2024_tot + v_duct2_2025_tot)
    ]
    
    for r_idx, data in enumerate(row_data_s2, start=17):
        for c_idx, val in enumerate(data, start=2):
            cell = ws_res.cell(row=r_idx, column=c_idx, value=val)
            cell.font = cell_font
            cell.border = thin_border
            if c_idx == 2:
                cell.alignment = Alignment(horizontal="left")
            elif c_idx == 3:
                cell.number_format = '#,##0.0'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="right")
                
    tot_row_s2 = 22
    ws_res.cell(row=tot_row_s2, column=2, value="TOTAL LIQUIDACIÓN (ESCENARIO 2)").font = total_font
    ws_res.cell(row=tot_row_s2, column=2).fill = subtotal_fill
    ws_res.cell(row=tot_row_s2, column=2).border = double_bottom_border
    
    ws_res.cell(row=tot_row_s2, column=3, value=f"=SUM(C17:C21)").font = total_font
    ws_res.cell(row=tot_row_s2, column=3).number_format = '#,##0.0'
    ws_res.cell(row=tot_row_s2, column=3).fill = subtotal_fill
    ws_res.cell(row=tot_row_s2, column=3).border = double_bottom_border
    
    for c_idx, col_let in enumerate(['D', 'E', 'F'], start=4):
        cell = ws_res.cell(row=tot_row_s2, column=c_idx, value=f"=SUM({col_let}17:{col_let}21)")
        cell.font = total_font
        cell.number_format = '$#,##0.00'
        cell.fill = subtotal_fill
        cell.border = double_bottom_border
        
    # Autoajustar columnas en Resumen
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        max_len = 0
        for row in range(5, 23):
            val = ws_res.cell(row=row, column=col).value
            if val is not None:
                if str(val).startswith('='):
                    max_len = max(max_len, 15)
                else:
                    max_len = max(max_len, len(str(val)))
        ws_res.column_dimensions[col_letter].width = max(max_len + 4, 18)
        
    wb.save(out_file)
    print(f"Liquidación final generada exitosamente en '{out_file}' con todos los datos y la hoja resumen.")

if __name__ == "__main__":
    main()
